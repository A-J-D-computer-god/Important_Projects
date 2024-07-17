try:
    import sys
    import openpyxl
    import pandas as pd
    import matplotlib.pyplot as plt
    import squarify
    from PyQt5.QtWidgets import QApplication, QMainWindow, QPushButton, QLineEdit, QTextEdit, QVBoxLayout, QWidget, QLabel, QComboBox, QMessageBox


    class ExcelApp(QMainWindow):
        def __init__(self):
            super().__init__()

            self.setWindowTitle("Excel Data Reader")
            self.setGeometry(100, 100, 800, 600)

            self.layout = QVBoxLayout()

            self.path_label = QLabel("Please Enter The path:")
            self.layout.addWidget(self.path_label)
            self.path_input = QLineEdit()
            self.layout.addWidget(self.path_input)

            self.duty_label = QLabel("On Duty:")
            self.layout.addWidget(self.duty_label)
            self.duty_input = QComboBox()
            self.duty_input.addItems(["Search", "Setande", "Masrafe Vasete", "Arzesh Afzoode Tafazol", "Arzesh Afzoode Jam"])
            self.layout.addWidget(self.duty_input)

            self.search_label = QLabel("Char/Year:")
            self.layout.addWidget(self.search_label)
            self.search_input = QLineEdit()
            self.layout.addWidget(self.search_input)

            self.output = QTextEdit()
            self.layout.addWidget(self.output)

            self.button = QPushButton("Answer")
            self.button.clicked.connect(self.execute)
            self.layout.addWidget(self.button)

            container = QWidget()
            container.setLayout(self.layout)
            self.setCentralWidget(container)

        def load_worksheet(self, path):
            wb = openpyxl.load_workbook(path)
            return wb.active

        def datafinder(self, ws, searchdata):
            for i in range(1, ws.max_row + 1):
                for j in range(1, ws.max_column + 1):
                    if searchdata == str(ws.cell(i, j).value):
                        return i, j
            return None

        def calculate_sum(self, df, column, indices):
            return pd.to_numeric(df.loc[indices, column], errors='coerce').fillna(0).sum()

        def execute(self):
            path = self.path_input.text()
            ws = self.load_worksheet(path)
            df = pd.read_excel(path)
            duty = self.duty_input.currentText().lower()
            searchdata = self.search_input.text()

            if duty == "search":
                result = self.datafinder(ws, searchdata)
                if result:
                    self.output.setText(f"Found: {searchdata} at Row: {result[0]}, Column: {result[1]}")
                else:
                    self.output.setText(f"{searchdata} not found.")

            elif duty == "setande":
                column = searchdata
                indices = [6, 7, 8, 9, 10, 11, 37]
                result_setade = self.calculate_sum(df, column, indices)
                self.output.setText(f"Sum: {result_setade}")
                self.plot_graph(df, column, result_setade, "Setande")

            elif duty == "masrafe vasete":
                column = searchdata
                indices = [14, 15, 16, 17, 19, 22, 23, 24, 25, 27, 30, 31, 32, 33, 35, 40, 56, 58]
                result_masrafe_vasete = self.calculate_sum(df, column, indices)
                self.output.setText(f"Sum: {result_masrafe_vasete}")
                self.plot_graph(df, column, result_masrafe_vasete, "Masrafe Vasete")

            elif duty == "arzesh afzoode tafazol":
                column = searchdata
                indices_setade = [6, 7, 8, 9, 10, 11, 37]
                result_setade = self.calculate_sum(df, column, indices_setade)
                indices_masrafe_vasete = [14, 15, 16, 17, 19, 22, 23, 24, 25, 27, 30, 31, 32, 33, 35, 40, 56, 58]
                result_masrafe_vasete = self.calculate_sum(df, column, indices_masrafe_vasete)
                result_arfz_tafazol = result_setade - result_masrafe_vasete
                self.output.setText(f"Sum: {result_arfz_tafazol}")
                self.plot_graph(df, column, result_arfz_tafazol, "Arzesh Afzoode Tafazol")

            elif duty == "arzesh afzoode jam":
                column = searchdata
                indices_positive = [13, 18, 21, 26, 29, 34, 39, 41, 42, 43, 57, 60]
                indices_negative = [45, 46, 47, 48, 49, 50, 51, 52, 53, 54]
                positive_sum = self.calculate_sum(df, column, indices_positive)
                negative_sum = self.calculate_sum(df, column, indices_negative)
                result = positive_sum - negative_sum
                self.output.setText(f"Sum: {result}")
                self.plot_graph(df, column, result, "Arzesh Afzoode Jam")

        def plot_graph(self, df, column, result, title):
            indices_setade = [6, 7, 8, 9, 10, 11, 37]
            result_setade = self.calculate_sum(df, column, indices_setade)
            indices_masrafe_vasete = [14, 15, 16, 17, 19, 22, 23, 24, 25, 27, 30, 31, 32, 33, 35, 40, 56, 58]
            result_masrafe_vasete = self.calculate_sum(df, column, indices_masrafe_vasete)
            sizes = [result_setade, result_masrafe_vasete, result_setade - result_masrafe_vasete]
            labels = ['Setande', 'Masaref Vasete', 'Arzesh Afzoode']
            colors = plt.cm.viridis([i / len(sizes) for i in range(len(sizes))])
            squarify.plot(sizes=sizes, label=labels, color=colors, alpha=.7)
            plt.axis('off')
            plt.title(title)
            plt.show()


    def main():
        app = QApplication(sys.argv)
        window = ExcelApp()
        window.show()
        sys.exit(app.exec_())

    
    if True:
        main()
except Exception as E:
    print(f"Just Some Problem : {E}")
