import sys
import openpyxl
import pandas as pd
import matplotlib.pyplot as plt
import squarify
from PyQt5.QtWidgets import (QApplication, QMainWindow, QPushButton, QLineEdit, QTextEdit, QGridLayout,
                             QWidget, QLabel, QComboBox, QMessageBox, QFileDialog)
from PyQt5.QtGui import QIcon, QFont
from PyQt5.QtCore import Qt, QSize


class ExcelApp(QMainWindow):
    def __init__(self):
        super().__init__()

        self.setWindowTitle("Excel Data Reader")
        self.setGeometry(100, 100, 800, 600)

        self.layout = QGridLayout()

        self.path_label = QLabel("Please Enter The Path:")
        self.path_label.setFont(QFont('Arial', 12))
        self.path_label.setStyleSheet("color: #ecf0f1;")
        self.layout.addWidget(self.path_label, 0, 0)

        self.path_input = QLineEdit()
        self.path_input.setStyleSheet("background-color: #34495e; color: #ecf0f1; padding: 5px; border-radius: 5px;")
        self.layout.addWidget(self.path_input, 0, 1)

        self.browse_button = QPushButton()
        self.browse_button.setIcon(QIcon('icon2.ico'))
        self.browse_button.setIconSize(QSize(24, 24))
        self.browse_button.clicked.connect(self.browse_file)
        self.layout.addWidget(self.browse_button, 0, 2)

        self.duty_label = QLabel("On Duty:")
        self.duty_label.setFont(QFont('Arial', 12))
        self.duty_label.setStyleSheet("color: #ecf0f1;")
        self.layout.addWidget(self.duty_label, 1, 0)

        self.duty_input = QComboBox()
        self.duty_input.addItems([
            "Search", "Setande", "Masrafe Vasete", "Arzesh Afzoode Tafazol", "Arzesh Afzoode Jam",
            "Compare(Setande)", "Compare(Masrafe Vasete)", "Compare(Arzesh Afzoode)"
        ])
        self.duty_input.setStyleSheet("background-color: #34495e; color: #ecf0f1; padding: 5px; border-radius: 5px;")
        self.layout.addWidget(self.duty_input, 1, 1)

        self.years_label = QLabel("Which Years? (In Compare Only):")
        self.years_label.setFont(QFont('Arial', 12))
        self.years_label.setStyleSheet("color: #ecf0f1;")
        self.layout.addWidget(self.years_label, 2, 0)

        self.years_input = QLineEdit()
        self.years_input.setStyleSheet("background-color: #34495e; color: #ecf0f1; padding: 5px; border-radius: 5px;")
        self.layout.addWidget(self.years_input, 2, 1)

        self.search_label = QLabel("Char/Year:")
        self.search_label.setFont(QFont('Arial', 12))
        self.search_label.setStyleSheet("color: #ecf0f1;")
        self.layout.addWidget(self.search_label, 3, 0)

        self.search_input = QLineEdit()
        self.search_input.setStyleSheet("background-color: #34495e; color: #ecf0f1; padding: 5px; border-radius: 5px;")
        self.layout.addWidget(self.search_input, 3, 1)

        self.output = QTextEdit()
        self.output.setStyleSheet("background-color: #34495e; color: #ecf0f1; padding: 10px; border-radius: 5px;")
        self.layout.addWidget(self.output, 4, 0, 1, 3)

        self.button = QPushButton("Answer")
        self.button.setIcon(QIcon('path/to/answer_icon.ico'))  # Add a path to an icon file
        self.button.setIconSize(QSize(24, 24))
        self.button.clicked.connect(self.execute)
        self.button.setStyleSheet(
            "background-color: #e74c3c; color: #ecf0f1; padding: 10px; border: none; border-radius: 5px;")
        self.layout.addWidget(self.button, 5, 0, 1, 3)

        container = QWidget()
        container.setLayout(self.layout)
        self.setCentralWidget(container)

        self.setStyleSheet("""
            QMainWindow {
                background-color: #2c3e50;
            }
            QLabel, QComboBox {
                font-size: 14px;
                padding: 5px;
            }
        """)

    def browse_file(self):
        file_name, _ = QFileDialog.getOpenFileName(self, "Open Excel File", "", "Excel Files (*.xlsx *.xls)")
        if file_name:
            self.path_input.setText(file_name)

    def load_worksheet(self, path):
        try:
            wb = openpyxl.load_workbook(path)
            return wb.active
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to load worksheet: {e}")
            return None

    def datafinder(self, ws, searchdata):
        for i in range(1, ws.max_row + 1):
            for j in range(1, ws.max_column + 1):
                if searchdata == str(ws.cell(i, j).value):
                    return i, j
        return None

    def calculate_sum(self, df, column, indices):
        return pd.to_numeric(df.loc[indices, column], errors='coerce').fillna(0).sum()

    def execute(self):
        path = self.path_input.text()
        ws = self.load_worksheet(path)
        if not ws:
            return

        try:
            df = pd.read_excel(path)
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to load Excel file: {e}")
            return

        duty = self.duty_input.currentText().lower()
        searchdata = self.search_input.text()

        if duty == "search":
            result = self.datafinder(ws, searchdata)
            if result:
                self.output.setText(f"Found: {searchdata} at Row: {result[0]}, Column: {result[1]}")
            else:
                self.output.setText(f"{searchdata} not found.")

        elif duty == "setande":
            column = searchdata
            indices = [6, 7, 8, 9, 10, 11, 37]
            result_setade = self.calculate_sum(df, column, indices)
            self.output.setText(f"Sum: {result_setade}")
            self.plot_graph(df, column, result_setade, "Setande")

        elif duty == "masrafe vasete":
            column = searchdata
            indices = [14, 15, 16, 17, 19, 22, 23, 24, 25, 27, 30, 31, 32, 33, 35, 40, 56, 58]
            result_masrafe_vasete = self.calculate_sum(df, column, indices)
            indices = [6, 7, 8, 9, 10, 11, 37]
            result_setade = self.calculate_sum(df, column, indices)
            percent = result_masrafe_vasete*100 / result_setade
            self.output.setText(f"Sum: {result_masrafe_vasete},{percent}")
            self.plot_graph(df, column, result_masrafe_vasete, "Masrafe Vasete")

        elif duty == "arzesh afzoode tafazol":
            column = searchdata
            indices_setade = [6, 7, 8, 9, 10, 11, 37]
            result_setade = self.calculate_sum(df, column, indices_setade)
            indices_masrafe_vasete = [14, 15, 16, 17, 19, 22, 23, 24, 25, 27, 30, 31, 32, 33, 35, 40, 56, 58]
            result_masrafe_vasete = self.calculate_sum(df, column, indices_masrafe_vasete)
            result_arfz_tafazol = result_setade - result_masrafe_vasete
            percent = result_arfz_tafazol*100 / result_setade
            self.output.setText(f"Sum: {result_arfz_tafazol},{percent}")
            self.plot_graph(df, column, result_arfz_tafazol, "Arzesh Afzoode Tafazol")

        elif duty == "arzesh afzoode jam":
            column = searchdata
            indices_positive = [13, 18, 21, 26, 29, 34, 39, 41, 42, 43, 57, 60]
            indices_negative = [45, 46, 47, 48, 49, 50, 51, 52, 53, 54]
            positive_sum = self.calculate_sum(df, column, indices_positive)
            negative_sum = self.calculate_sum(df, column, indices_negative)
            result = positive_sum - negative_sum
            indices_setade = [6, 7, 8, 9, 10, 11, 37]
            result_setade = self.calculate_sum(df, column, indices_setade)
            percent = result*100 / result_setade
            self.output.setText(f"Sum: {result},{percent}")
            self.plot_graph(df, column, result, "Arzesh Afzoode Jam")

        elif duty == "compare(setande)":
            years = self.years_input.text().split(" ")
            labels = years
            values = []
            for year in years:
                if year in df.columns:
                    indices = [6, 7, 8, 9, 10, 11, 37]
                    result_setade = self.calculate_sum(df, year, indices)
                    values.append(result_setade)
                else:
                    values.append(0)
                    QMessageBox.warning(self, "Warning", f"Column '{year}' not found in the data.")

            plt.bar(labels, values)
            plt.xlabel('Year')
            plt.ylabel('Setande')
            plt.title('Setande Compare')
            plt.show()

        elif duty == "compare(masrafe vasete)":
            years = self.years_input.text().split(" ")
            labels = years
            values = []
            for year in years:
                if year in df.columns:
                    indices = [14, 15, 16, 17, 19, 22, 23, 24, 25, 27, 30, 31, 32, 33, 35, 40, 56, 58]
                    result_Masrafe_Vasete = self.calculate_sum(df, year, indices)
                    values.append(result_Masrafe_Vasete)
                else:
                    values.append(0)
                    QMessageBox.warning(self, "Warning", f"Column '{year}' not found in the data.")

            plt.bar(labels, values)
            plt.xlabel('Year')
            plt.ylabel('Masrafe Vasete')
            plt.title('Masrafe Vasete Compare')
            plt.show()

        elif duty == "compare(arzesh afzoode)":
            years = self.years_input.text().split(" ")
            labels = years
            values = []
            for year in years:
                if year in df.columns:
                    column = searchdata
                    indices = [6, 7, 8, 9, 10, 11, 37]
                    result_setade = self.calculate_sum(df, column, indices)
                    indices = [14, 15, 16, 17, 19, 22, 23, 24, 25, 27, 30, 31, 32, 33, 35, 40, 56, 58]
                    result_Masrafe_Vasete = self.calculate_sum(df, year, indices)
                    result_Arzesh_Afzoode = result_setade - result_Masrafe_Vasete
                    values.append(result_Arzesh_Afzoode)
                else:
                    values.append(0)
                    QMessageBox.warning(self, "Warning", f"Column '{year}' not found in the data.")

            plt.bar(labels, values)
            plt.xlabel('Year')
            plt.ylabel('Arzesh Afzoode')
            plt.title('Arzesh Afzoode Compare')
            plt.show()

    def plot_graph(self, df, column, result, title):
        indices_setade = [6, 7, 8, 9, 10, 11, 37]
        result_setade = self.calculate_sum(df, column, indices_setade)
        indices_masrafe_vasete = [14, 15, 16, 17, 19, 22, 23, 24, 25, 27, 30, 31, 32, 33, 35, 40, 56, 58]
        result_masrafe_vasete = self.calculate_sum(df, column, indices_masrafe_vasete)
        sizes = [result_setade, result_masrafe_vasete, result_setade - result_masrafe_vasete]
        labels = ['Setande', 'Masaref Vasete', 'Arzesh Afzoode']
        colors = plt.cm.viridis([i / len(sizes) for i in range(len(sizes))])
        squarify.plot(sizes=sizes, label=labels, color=colors, alpha=.7)
        plt.axis('off')
        plt.title(title)
        plt.show()


def main():
    app = QApplication(sys.argv)
    window = ExcelApp()
    window.show()
    sys.exit(app.exec_())


if __name__ == "__main__":
    main()
