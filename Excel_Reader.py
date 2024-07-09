import openpyxl
import pandas as pd


def load_worksheet(path):
    wb = openpyxl.load_workbook(path)
    return wb.active


def datafinder(ws, searchdata):
    for i in range(1, ws.max_row + 1):
        for j in range(1, ws.max_column + 1):
            if searchdata == str(ws.cell(i, j).value):
                print(f"Found: {ws.cell(i, j).value} at Row: {i}, Column: {j}")
                return i, j
    return None


def calculate_sum(df, column, indices):
    return pd.to_numeric(df.loc[indices, column], errors='coerce').fillna(0).sum()


def main():
    while True:
        try:
            path = input("Please Enter The path (or type 'exit' to quit): ")
            if path.lower() == 'exit':
                break

            ws = load_worksheet(path)
            df = pd.read_excel(path)

            duty = input(
                "On Duty... Search, Setande, Masrafe Vasete, Arzesh Afzoode Tafazol, Arzesh Afzoode Jam: ").lower()

            if duty == "search":
                searchdata = input("What Are You Looking For?: ")
                datafinder(ws, searchdata)

            elif duty == "setande":
                searchdata = input("Which Year?: ")
                datafinder(ws, searchdata)
                column = searchdata
                indices = [6, 7, 8, 9, 10, 11, 37]
                result_setade = calculate_sum(df, column, indices)
                print(f"Sum: {result_setade}")

            elif duty == "masrafe vasete":
                searchdata = input("Which Year?: ")
                datafinder(ws, searchdata)
                column = searchdata
                indices = [14, 15, 16, 17, 19, 22, 23, 24, 25, 27, 30, 31, 32, 33, 35, 40, 56, 58]
                result_masrafe_vasete = calculate_sum(df, column, indices)
                print(f"Sum: {result_masrafe_vasete}")

            elif duty == "arzesh afzoode tafazol":
                column = input("Which year?: ")
                indices_setade = [6, 7, 8, 9, 10, 11, 37]
                result_setade = calculate_sum(df, column, indices_setade)

                indices_masrafe_vasete = [14, 15, 16, 17, 19, 22, 23, 24, 25, 27, 30, 31, 32, 33, 35, 40, 56, 58]
                result_masrafe_vasete = calculate_sum(df, column, indices_masrafe_vasete)

                result_arfz_tafazol = result_setade - result_masrafe_vasete
                print(f"Sum: {result_arfz_tafazol}")

            elif duty == "arzesh afzoode jam":
                searchdata = input("Which Year?: ")
                datafinder(ws, searchdata)
                column = searchdata
                indices_positive = [13, 18, 21, 26, 29, 34, 39, 41, 42, 43, 57, 60]
                indices_negative = [45, 46, 47, 48, 49, 50, 51, 52, 53, 54]
                positive_sum = calculate_sum(df, column, indices_positive)
                negative_sum = calculate_sum(df, column, indices_negative)
                result = positive_sum - negative_sum
                print(f"Sum: {result}")

            else:
                print("Invalid duty selected. Please try again.")

        except Exception as e:
            print(f"An error occurred: {e}")


if __name__ == "__main__":
    main()
